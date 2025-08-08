from django.shortcuts import render,redirect, get_object_or_404
from .forms import RegisterForm, AddProductForm, ProductDetailForm,VehicleForm, ShopForm,ProcessReturnForm,ProcessDeliveryForm
from .forms import SalesTransactionItemReturnForm,AddItemToSaleForm,FinalizeSaleForm,SalesTransactionItemReturnFormSet # Import the formset
from .forms import  AddStockForm #,UpdatePaymentTypeForm
from django.contrib.auth import login, logout
from .utils import authenticate 
from django.contrib.auth.models import User 
from django.contrib.auth.decorators import login_required
from .models import AddProduct, ProductDetail,Sale, Vehicle, Shop, SalesTransaction, SalesTransactionItem
from accounts.models import ShopFinancialTransaction  # model from account app
from claim.models import Claim
from gov_agency.models import AdminProfile
from django.contrib import messages
from django.db.models  import Q,ProtectedError
from django.db import transaction,IntegrityError# For atomic operations
from django.http import JsonResponse
from decimal import Decimal
from django.db.models import Sum,Count,F, ExpressionWrapper, fields, DecimalField
from django.utils import timezone
from datetime import timedelta, date
import json 
from django.db.models.functions import TruncMonth,TruncDay
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.urls import resolve
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from expense.models import Expense
from django.db.models.functions import TruncDate,TruncMonth
from collections import defaultdict
from gov_agency.decorators import admin_mode_required
from django import forms
from itertools import groupby
from operator import attrgetter
from django.views.decorators.http import require_POST


# login view.........
def user_login(request):
    if request.method == 'POST':
        print("inside")
        email = request.POST.get("email")
        password = request.POST.get("password")
        user = authenticate(email=email, password=password)
        if user is not None:
            login(request, user)
            print("success............................................................")
            return redirect('dashboard:main_dashboard')  # Your custom dashboard view
        else:
            print("authantication is failed...............................................")
    return render(request, 'registration/login.html')
            
# sign up view ............
def register_user(request):
    print("inside custom register view..................")
    if request.user.is_authenticated:
        return redirect('dashboard:main_dashboard')
    
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request,user)
            messages.success(request, "Registration successful! You are now logged in.")
            return redirect('dashboard:main_dashboard')
    else:
        form = RegisterForm()    
    return render(request, 'registration/sign_up.html', {'form': form})



# add product view .......................................................................................................
@login_required
def create_product(request):
    form_in_modal_has_errors = False  # Flag to tell template to open modal on page load if errors exist

    if request.method == 'POST':
        form = AddProductForm(request.POST)
        if form.is_valid():
            product_instance = form.save(commit=False)
            product_instance.user = request.user 
            try:
                product_instance.save() 
                messages.success(request, f"Product '{product_instance.name}' added successfully.")
                return redirect('stock:create_product') 
            except IntegrityError:
                messages.error(request, f"A product named '{product_instance.name}' already exists for your account.")
                form_in_modal_has_errors = True
            except Exception as e:
                messages.error(request, f"Could not save product. An error occurred: {str(e)}")
                form_in_modal_has_errors = True
        else:
            messages.error(request, "Please correct the errors indicated in the form.")
            form_in_modal_has_errors = True
    else:
        form = AddProductForm()

    products_list = AddProduct.objects.filter(user=request.user).order_by('-created_at')
    context = {
        'form': form,
        'products': products_list,
        'form_in_modal_has_errors': form_in_modal_has_errors,
    }
    return render(request, 'stock/create_product.html', context)

@login_required
@admin_mode_required
def confirm_delete_product_view(request, product_id):
    product = get_object_or_404(AddProduct, id=product_id, user=request.user) # Ensure user ownership
    context = {
        'product': product
    }
    return render(request, 'stock/confirm_delete_product.html', context)


@login_required
@admin_mode_required
def delete_product_view(request, product_id):
    product_to_delete = get_object_or_404(AddProduct, id=product_id, user=request.user)
    if request.method == 'POST':
        product_name = product_to_delete.name
        product_to_delete.delete()
        messages.error(request, f"Product '{product_name}' deleted successfully.")
        return redirect('stock:create_product')
    return redirect('stock:confirm_delete_product', product_id=product_id)

# product detail views .........................................................................................................................
@login_required
def add_product_details(request):
    form_had_errors = False
    query = request.GET.get('q','')

    if request.method == 'POST':
        # This POST is for adding a new ProductDetail via the modal
        form = ProductDetailForm(request.POST, user = request.user)
        if form.is_valid():
            detail = form.save(commit=False)
            detail.user = request.user
            detail.save()
            messages.success(request, f"Details for '{detail.product_base.name}' added successfully!")
            redirect_url = redirect('stock:add_product_details').url
            if query:
                redirect_url += f'?q={query}'
            return redirect(redirect_url)
        else:
            messages.error(request, "Please correct the errors in the form.")
            form_had_errors = True
    else:
        form = ProductDetailForm(user = request.user)

    # Fetch product details for the current user   
    product_details_qs = ProductDetail.objects.filter(user=request.user).select_related('product_base').order_by('-created_at')

    if query:
        product_details_qs = product_details_qs.filter(
            Q(product_base__name__icontains=query)|
            Q(packing_type__icontains=query)
        )

    # Check if the current user has any base products to select from for the form
    user_has_base_products =  AddProduct.objects.filter(user=request.user).order_by('-created_at')
    if not user_has_base_products and request.method == 'GET' and not query:
        messages.warning(request, "You need to create a base product first (in 'Add Products') before adding details.")
    

    context = {
        'form': form,
        'product_details': product_details_qs,
        'form_had_errors': form_had_errors,
        'search_query': query,
        'user_has_base_products': user_has_base_products,
    }
    return render(request, 'stock/add_product_details.html', context)



@login_required
@admin_mode_required
def product_detail_delete_selected_view(request):
    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_details_ids')
        products_to_delete = ProductDetail.objects.filter(id__in=selected_ids, user=request.user)
        if selected_ids:
            try:
                ProductDetail.objects.filter(id__in=selected_ids, user=request.user).delete()
                messages.success(request, f"{len(selected_ids)} product detail(s) deleted successfully.")
            except ProtectedError as e:
                protected_sales_items = e.protected_objects
                protected_transaction_ids = {item.transaction.id for item in protected_sales_items}
                conflicting_sales = SalesTransaction.objects.filter(id__in=protected_transaction_ids).distinct()
                messages.error(request, "Deletion failed because these products are part of existing sales records. Please review and handle these sales first.")
                context = {
                    'products_to_delete': products_to_delete,
                    'conflicting_sales': conflicting_sales,
                }
                return render(request, 'stock/deletion_blocked_by_sales.html', context)
    return redirect('stock:add_product_details')

@login_required
@admin_mode_required
def product_detail_update_view(request, pk):
    instance  = get_object_or_404(ProductDetail, pk=pk, user=request.user)
    if request.method == 'POST':
        form = ProductDetailForm(request.POST, instance=instance , user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, f"Details for '{instance .product_base.name}' updated successfully!")
            query = request.GET.get('q_after_update', '')
            redirect_url = redirect('stock:add_product_details').url
            if query: redirect_url += f'?q={query}'
            return redirect(redirect_url)
        else:
            messages.error(request, "Error updating. Please check the data and try again.")
            return redirect('stock:add_product_details')
    else:
        return redirect('stock:add_product_details')
    

@login_required
def add_stock_to_product_detail_view(request, pk):
    # Get the ProductDetail instance, ensuring user ownership
    product_detail = get_object_or_404(ProductDetail, pk=pk, user=request.user)
    
    if request.method == 'POST':
        form = AddStockForm(request.POST, product_detail_instance=product_detail)
        if form.is_valid():
            new_stock_quantity = form.cleaned_data['new_stock_quantity']
            new_expiry_date = form.cleaned_data['new_expiry_date']

            try:
                with transaction.atomic():
                    # Check if a batch with the exact same new expiry date already exists for this product
                    existing_batch =  ProductDetail.objects.select_for_update().get(pk=pk, user=request.user)
                    if existing_batch:
                        # Add stock to the existing batch with the same expiry date
                        existing_batch.increase_stock(new_stock_quantity)   
                        existing_batch.expirey_date = new_expiry_date
                        existing_batch.save(update_fields=['stock', 'expirey_date', 'updated_at'])
                        messages.success(request, f"Added {new_stock_quantity} stock to existing batch of {existing_batch.product_base.name} (Exp: {new_expiry_date}). New stock: {existing_batch.stock}.")
                    else:
                        # Create a new ProductDetail instance for the new batch
                        new_batch = ProductDetail.objects.create(
                            product_base=product_detail.product_base,
                            user=request.user,
                            packing_type=product_detail.packing_type,
                            quantity_in_packing=product_detail.quantity_in_packing,
                            unit_of_measure=product_detail.unit_of_measure,
                            items_per_master_unit=product_detail.items_per_master_unit,
                            price_per_item=product_detail.price_per_item,
                            stock=new_stock_quantity,
                            expirey_date=new_expiry_date
                        )
                        messages.success(request, f"New stock batch created for {new_batch.product_base.name} with quantity {new_stock_quantity} and expiry {new_expiry_date}.")
                    
                return redirect('stock:add_product_details')
            except Exception as e:
                messages.error(request, f"An error occurred while adding stock: {str(e)}")
        else:
            # This is tricky for a modal. We redirect with a generic error.
            # A better UX would use AJAX for this form submission.
            error_string = ". ".join([f"{field}: {', '.join(errors)}" for field, errors in form.errors.items()])
            messages.error(request, f"Failed to add stock. {error_string}")
            return redirect('stock:add_product_details')
    
    # This view is for POST only from the modal
    return redirect('stock:add_product_details')


#  sales views...................................................................................................................

@login_required
def sales_processing_view(request):
    """
    Handles the entire multi-item sales process, including adding items to a
    session-based cart and finalizing the transaction with an overall discount.
    """
    # Session key for the current "cart" of sale items
    current_transaction_items_session_key = f'current_transaction_items_{request.user.id}'
    
    # Initialize forms for both GET and POST requests
    add_item_form = AddItemToSaleForm(user=request.user)
    finalize_sale_form = FinalizeSaleForm(user=request.user)

    if request.method == 'POST':
        action = request.POST.get('action')

        # --- ACTION 1: ADD ITEM TO THE CURRENT TRANSACTION ---
        if action == 'add_item_to_transaction':
            add_item_form = AddItemToSaleForm(request.POST, user=request.user)
            if add_item_form.is_valid():
                product_detail_batch = add_item_form.cleaned_data['product_detail_batch']
                quantity_to_add = add_item_form.cleaned_data['quantity_to_add']
                selling_price_item = add_item_form.cleaned_data['selling_price_per_item']
                current_items = request.session.get(current_transaction_items_session_key, [])

                # Check for duplicates
                if any(item['product_detail_id'] == product_detail_batch.id for item in current_items):
                    messages.warning(request, f"{product_detail_batch.product_base.name} is already in the list. Remove to re-add with new quantity/price.")
                else:
                    # Calculate gross subtotal for this line (no discount here)
                    individual_items_count = product_detail_batch._get_items_from_decimal(quantity_to_add)
                    gross_subtotal = individual_items_count * selling_price_item
                    
                    current_items.append({
                        'product_detail_id': product_detail_batch.id,
                        'product_display_name': f"{product_detail_batch.product_base.name} {product_detail_batch.quantity_in_packing} {product_detail_batch.unit_of_measure} (Exp: {product_detail_batch.expirey_date.strftime('%d-%b-%Y')})",
                        'quantity_decimal': str(quantity_to_add),
                        'selling_price_per_item': str(selling_price_item),
                        'cost_price_per_item': str(product_detail_batch.price_per_item),
                        'line_subtotal': str(gross_subtotal.quantize(Decimal('0.01'))) # Store the gross (undiscounted) subtotal
                    })
                    messages.success(request, f"Added {quantity_to_add} of {product_detail_batch.product_base.name} to transaction.")
                
                request.session[current_transaction_items_session_key] = current_items
            return redirect('stock:sales')

        # --- ACTION 2: REMOVE ITEM FROM THE CURRENT TRANSACTION ---
        elif action == 'remove_item_from_transaction':
            item_index_to_remove = request.POST.get('item_index')
            current_items = request.session.get(current_transaction_items_session_key, [])
            try:
                item_index = int(item_index_to_remove)
                if 0 <= item_index < len(current_items):
                    current_items.pop(item_index)
                    request.session[current_transaction_items_session_key] = current_items
                    messages.info(request, "Item removed from transaction.")
                else:
                    messages.error(request, "Invalid item index to remove.")
            except (ValueError, TypeError):
                messages.error(request, "Error removing item.")
            return redirect('stock:sales')

        # --- ACTION 3: FINALIZE THE ENTIRE TRANSACTION ---
        elif action == 'finalize_transaction':
            current_items = request.session.get(current_transaction_items_session_key, [])
            if not current_items:
                messages.warning(request, "Cannot complete an empty transaction.")
                return redirect('stock:sales')

            # We use a new variable name here to avoid confusion with the empty form instance
            form_to_validate = FinalizeSaleForm(request.POST, user=request.user)
            if form_to_validate.is_valid():
                try:
                    with transaction.atomic():
                        # 1. CALCULATE GRAND TOTAL FROM SESSION DATA
                        gross_subtotal = sum(Decimal(item['line_subtotal']) for item in current_items)
                        discount = form_to_validate.cleaned_data.get('total_discount_amount') or Decimal('0.00')
                        grand_total = (gross_subtotal - discount).quantize(Decimal('0.01'))

                        # 2. GET ALL PAYMENT-RELATED DATA FROM THE FORM
                        payment_type = form_to_validate.cleaned_data.get('payment_type')
                        cash = form_to_validate.cleaned_data.get('amount_paid_cash') or Decimal('0.00')
                        online = form_to_validate.cleaned_data.get('amount_paid_online') or Decimal('0.00')
                        credit = form_to_validate.cleaned_data.get('amount_on_credit') or Decimal('0.00')

                        # 3. CREATE THE TRANSACTION INSTANCE (but don't save yet)
                        sales_transaction_header = form_to_validate.save(commit=False)
                        sales_transaction_header.user = request.user
                        sales_transaction_header.status = 'PENDING_DELIVERY' if sales_transaction_header.needs_vehicle else 'COMPLETED'
                        
                        # 4. HANDLE PAYMENT LOGIC (Single vs. Split)
                        if payment_type == 'SPLIT':
                            # For split payments, validate that the sum matches the grand total
                            if (cash + online + credit).quantize(Decimal('0.01')) != grand_total:
                                raise forms.ValidationError(f"Split payments (Rs {cash + online + credit}) do not match Grand Total (Rs {grand_total}).")
                            
                            sales_transaction_header.notes = (
                            f"Split Payment: Cash = Rs {cash:.2f}, "
                            f"Online = Rs {online:.2f}, "
                            f"Credit = Rs {credit:.2f}"
                            )
                            # The amounts from the form are already correct on the instance
                        else:
                            # For single payments, we override the amount fields
                            sales_transaction_header.amount_paid_cash = grand_total if payment_type == 'CASH' else Decimal('0.00')
                            sales_transaction_header.amount_paid_online = grand_total if payment_type == 'ONLINE' else Decimal('0.00')
                            sales_transaction_header.amount_on_credit = grand_total if payment_type == 'CREDIT' else Decimal('0.00')
                        
                        # Set grand totals manually before saving
                        sales_transaction_header.grand_total_revenue = grand_total
                        # grand_total_cost will be calculated later

                        sales_transaction_header.save() # First save to get a PK

                        # 5. CREATE SALE ITEMS AND DECREASE STOCK (Unchanged logic)
                        for item_data in current_items:
                            pd_batch = ProductDetail.objects.select_for_update().get(pk=item_data['product_detail_id'])
                            if not pd_batch.decrease_stock(Decimal(item_data['quantity_decimal'])):
                                raise Exception(f"Stock update failed for {pd_batch}. Transaction rolled back.")
                            SalesTransactionItem.objects.create(
                                transaction=sales_transaction_header,
                                product_detail_snapshot=pd_batch,
                                quantity_sold_decimal=Decimal(item_data['quantity_decimal']),
                                selling_price_per_item=Decimal(item_data['selling_price_per_item']),
                                cost_price_per_item_at_sale=Decimal(item_data['cost_price_per_item']),
                            )
                        
                        # 6. UPDATE GRAND TOTALS (especially grand_total_cost) AND CREATE LEDGER ENTRY
                        sales_transaction_header.update_grand_totals() # This will calculate and save the cost
                        
                        # Create a ledger entry if there is a credit amount
                        if sales_transaction_header.amount_on_credit > 0:
                            ShopFinancialTransaction.objects.create(
                                shop=sales_transaction_header.customer_shop,
                                customer_name_snapshot=sales_transaction_header.customer_name_manual,
                                user=request.user,
                                source_sale=sales_transaction_header,
                                transaction_type='CREDIT_SALE',
                                debit_amount=sales_transaction_header.amount_on_credit,
                                notes=f"Credit from Sale Transaction #{sales_transaction_header.pk}"
                            )
                        
                        messages.success(request, f"Transaction #{sales_transaction_header.pk} completed successfully!")
                        del request.session[current_transaction_items_session_key] # Clear the cart
                        return redirect('stock:sale_receipt', sale_pk=sales_transaction_header.pk)

                except ProductDetail.DoesNotExist:
                    messages.error(request, "Error: A product in the transaction could not be found. Transaction cancelled.")
                except Exception as e:
                    messages.error(request, f"An unexpected error occurred: {str(e)}")
            else:
                messages.error(request, f"{form_to_validate.errors.as_text()}")

    # --- For GET requests or if a POST fails validation ---
    current_transaction_items = request.session.get(current_transaction_items_session_key, [])
    current_transaction_subtotal = sum(Decimal(item['line_subtotal']) for item in current_transaction_items)

    context = {
        'add_item_form': add_item_form,
        'finalize_sale_form': finalize_sale_form,
        'current_transaction_items': current_transaction_items,
        'current_transaction_subtotal': current_transaction_subtotal,
    }
    return render(request, 'stock/sales_multiem.html', context)




@login_required
def bulk_delete_sales_view(request):
    """
    Handles the deletion of a specific list of sales transactions,
    typically those that are blocking a product deletion.
    """
    if request.method == 'POST':
        sales_ids_str = request.POST.get('sales_ids_to_delete')
        if sales_ids_str:
            try:
                # Convert the string of IDs '1,2,3' into a list of integers
                sales_ids = [int(id_str) for id_str in sales_ids_str.split(',') if id_str.isdigit()]
                
                # Fetch the transactions belonging to the current user to ensure security
                transactions_to_delete = SalesTransaction.objects.filter(pk__in=sales_ids, user=request.user)
                
                count = transactions_to_delete.count()
                
                if count > 0:
                    transactions_to_delete.delete()
                    messages.success(request, f"Successfully deleted {count} conflicting sales record(s). You may now try deleting the product(s) again.")
                else:
                    messages.warning(request, "No matching sales records were found to delete.")

            except Exception as e:
                messages.error(request, f"An error occurred while deleting sales: {str(e)}")
        else:
            messages.warning(request, "No sales IDs were provided for deletion.")

    # Always redirect back to the product details page, as the user's task is now to re-attempt the product deletion.
    return redirect('stock:add_product_details')




# AJAX endpoint to get details for a selected product_detail_batch
# This is used by JavaScript to auto-fill parts of the SaleForm
@login_required
def ajax_get_batch_details_for_sale(request, pk): # Renamed from get_sale_product_batch_details_json
    try:
        detail = get_object_or_404(ProductDetail, pk=pk, user=request.user)
        
        # Ensure properties total_items_in_stock and display_stock are robust in ProductDetail model
        data = {
            'pk': detail.pk,
            'product_name_display': str(detail.product_base.name),
            'expiry_date_display': detail.expirey_date.strftime('%Y-%m-%d'),
            'current_stock_master_units_display': str(detail.stock), # e.g., "2.5"
            'current_stock_master_units_unit': "units" , # You might want a unit field on ProductDetail or AddProduct
            'total_individual_items_available': detail.total_items_in_stock,
            'cost_price_per_item': str(detail.price_per_item),
            'suggested_selling_price_per_item': str(round(detail.price_per_item * Decimal('1.20'), 2)), # 20% markup
            'items_per_master_unit': detail.items_per_master_unit,
        }
        return JsonResponse({'success': True, 'data': data})
    except ProductDetail.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Product batch not found or not authorized.'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'An error occurred in AJAX: {str(e)}'}, status=500)


@login_required
def all_transactions_list_view(request):
    transactions_list = SalesTransaction.objects.filter(
        user=request.user # Or remove/adjust if admins see all
    ).select_related(
        'customer_shop', 
        'assigned_vehicle'
    ).prefetch_related(
        'items__product_detail_snapshot__product_base' # For product summary
    ).order_by('-transaction_time')

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if start_date and not end_date:
        # **Special Case**: If only start_date is provided, filter for that single day.
        # The '__date' lookup matches the date part of a DateTimeField.
        transactions_list = transactions_list.filter(transaction_time__date=start_date)
    else:
        # Standard range filter: apply if dates are present.
        # This works if both are present, or if only end_date is present.
        if start_date:
            transactions_list = transactions_list.filter(transaction_time__date__gte=start_date)
        if end_date:
            transactions_list = transactions_list.filter(transaction_time__date__lte=end_date)
    # Pagination (optional, but good for long lists)
    paginator = Paginator(transactions_list, 50) # Show 50 transactions per page
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1) # If page is not an integer, deliver first page.
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        page_obj = paginator.page(paginator.num_pages)

    context = {
        'page_obj': page_obj, # Use page_obj in template for pagination
        'transactions': page_obj.object_list, # The transactions for the current page
    }
    return render(request, 'stock/all_transactions_list.html', context)


@login_required
def reverse_sale_prompt_view(request):
    if request.method == 'POST':
        sale_id = request.POST.get('sale_id')
        return redirect('stock:reverse_sale_confirm', sale_id=sale_id)
    return render(request, 'stock/reverse_sale_prompt.html')


@login_required
def get_sale_info(request, sale_id):
    try:
        sale = SalesTransaction.objects.get(pk=sale_id, user=request.user)
        return JsonResponse({
            "success": True,
            "date": sale.transaction_time.strftime("%Y-%m-%d")
        })
    except SalesTransaction.DoesNotExist:
        return JsonResponse({"success": False})

@login_required
@transaction.atomic
def confirm_reverse_sale(request, sale_id):
    sale = get_object_or_404(SalesTransaction, pk=sale_id, user=request.user)

    if request.method == "POST":
        try:
            if sale.status != "PENDING_DELIVERY":
                # 1. Restore Stock
                for item in sale.items.select_related("product_detail_snapshot"):
                    item.product_detail_snapshot.increase_stock(item.quantity_sold_decimal)

                # 2. Delete Ledger Entry if it was CREDIT
                if sale.payment_type == 'CREDIT' or sale.payment_type == 'SPLIT':
                    ShopFinancialTransaction.objects.filter(source_sale=sale).delete()

                # 3. Delete Sale Items and Sale
                sale.items.all().delete()
                sale.delete()

                messages.success(request, f"Sale #{sale_id} reversed and deleted successfully.")
                return redirect('stock:all_transactions_list')
            else:
                messages.warning(request,"Sale Should not be in pending state if you want ot reverse.")
        except Exception as e:
            messages.error(request, f"Error while reversing sale: {str(e)}")
            return redirect('stock:all_transactions_list')

    return redirect('stock:all_transactions_list')


@login_required
def export_sales_to_excel(request):
    """
    Handles the export of sales transactions to an Excel (.xlsx) file.
    Can be filtered by a date range OR a specific list of transaction IDs.
    """
    # Start with the base queryset for the logged-in user
    transactions_list = SalesTransaction.objects.filter(user=request.user)

    # --- NEW: Check for a specific list of IDs first ---
    ids_to_export_str = request.GET.get('ids')
    if ids_to_export_str:
        try:
            # Convert comma-separated string of IDs to a list of integers
            ids_list = [int(id_str) for id_str in ids_to_export_str.split(',') if id_str.isdigit()]
            if ids_list:
                transactions_list = transactions_list.filter(pk__in=ids_list)
        except (ValueError, TypeError):
            # If IDs are invalid, ignore and proceed to date filters
            pass
    else:
        # --- EXISTING DATE FILTER LOGIC ---
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        if start_date and not end_date:
            transactions_list = transactions_list.filter(transaction_time__date=start_date)
        else:
            if start_date:
                transactions_list = transactions_list.filter(transaction_time__date__gte=start_date)
            if end_date:
                transactions_list = transactions_list.filter(transaction_time__date__lte=end_date)

    # Apply optimizations and ordering
    transactions_list = transactions_list.select_related(
        'customer_shop', 'assigned_vehicle'
    ).prefetch_related( # Prefetch items and their nested details for performance
        'items__product_detail_snapshot__product_base'
    ).order_by('-transaction_time')

    # Create the Excel Workbook and Worksheet in memory
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # Define the Headers
    columns = [
        "Tx ID", "Date", "Time", "Customer", "Payment", "Status",
        "Total Revenue", "Total Profit", "Notes",  # Transaction-level info
        "Product Name", "Price Per Unit", "Quantity Sold", "Returned", "Discount", "Unit", # Item-level info
    ]
    ws.append(columns)

    # Style the header row (bold font)
    bold_font = Font(bold=True)
    for col_num in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = bold_font

    # Loop through the queryset and append data rows
    for tx in transactions_list:
        # Using .all() on a prefetched queryset is efficient (no extra DB hit here)
        all_items = tx.items.all()
        customer_name = tx.customer_shop.name if tx.customer_shop else tx.customer_name_manual or "N/A"

        # If a transaction has no items, we might still want to log it
        if not all_items:
            row = [
                tx.pk,
                tx.transaction_time.date(),
                tx.transaction_time.time().strftime('%H:%M:%S'),
                customer_name,
                tx.get_payment_type_display(),
                tx.get_status_display(),
                tx.grand_total_revenue, # Use number format for better sorting in Excel
                tx.calculated_grand_profit,
                tx.notes,
            ]
            ws.append(row)
        else:
            # If there are items, create a row for each one
            for item in all_items:
                row = [
                    tx.pk,
                    tx.transaction_time.date(),
                    tx.transaction_time.time().strftime('%H:%M:%S'),
                    customer_name,
                    tx.get_payment_type_display(),
                    tx.get_status_display(),
                    tx.grand_total_revenue,
                    tx.calculated_grand_profit,
                    tx.notes,
                    item.product_detail_snapshot.product_base.name,
                    item.selling_price_per_item,
                    item.quantity_sold_decimal,
                    item.returned_quantity_decimal,
                    f'{tx.total_discount_amount}', 
                    f'{item.product_detail_snapshot.quantity_in_packing} {item.product_detail_snapshot.unit_of_measure}',
                ]
                ws.append(row)

    # Auto-size columns for better readability
    for col_num, column_title in enumerate(columns, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].autosize = True

    # Prepare the HTTP response to serve the file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    # Define a dynamic filename
    filename = "sales_report.xlsx"
    if ids_to_export_str:
        filename = "conflicting_sales_export.xlsx"
    elif start_date and end_date:
        filename = f"sales_{start_date}_to_{end_date}.xlsx"
    elif start_date:
        filename = f"sales_{start_date}.xlsx"
        
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save the workbook to the response
    wb.save(response)

    return response




@login_required
def update_note(request):
    tx_id = request.POST.get('tx_id')
    new_note = request.POST.get('note', '')
    url = request.POST.get('next')
    tx = get_object_or_404(SalesTransaction, pk=tx_id)
    if tx:
        tx.notes = new_note
        tx.save()
    else:
        messages.error(request, f"no Transaction found")
    if url == 'shop_purchase_history':
        shop_pk = request.POST.get('shop_pk') 
        return redirect('stock:shop_purchase_history', shop_pk=shop_pk)
    return redirect('stock:all_transactions_list')  # or the same page






@login_required
def pending_deliveries_view(request):
    """
    UPDATED: This view now also calculates a consolidated "Loading Sheet" for each
    vehicle with pending deliveries.
    """
    # Get all pending transactions for the user, ordered by vehicle
    pending_transactions = SalesTransaction.objects.filter(
        user=request.user,
        status='PENDING_DELIVERY',
        assigned_vehicle__isnull=False # Only consider sales with an assigned vehicle
    ).select_related(
        'customer_shop', 'assigned_vehicle'
    ).prefetch_related(
        'items__product_detail_snapshot__product_base'
    ).order_by('-transaction_time')

    # --- CORRECTED AGGREGATION LOGIC ---
    # Structure: {vehicle_obj: {product_detail_obj: total_individual_item_count}}
    vehicle_loading_sheets_agg = defaultdict(lambda: defaultdict(int))
    no_of_invoices = defaultdict(lambda: int(0))  # To count invoices per vehicle
    for tx in pending_transactions:
        vehicle = tx.assigned_vehicle
        no_of_invoices[vehicle.pk] += 1
        for item in tx.items.all():
            product_detail = item.product_detail_snapshot
            # Get the quantity sold as a count of individual items
            individual_items_sold = item.dispatched_individual_items_count # Use the property
            # Add this integer count to the running total for that product on that vehicle
            vehicle_loading_sheets_agg[vehicle][product_detail] += individual_items_sold

    # Convert the aggregated individual item counts back to the decimal format for display
    # The final structure will be: {vehicle_object: [ {'product': product_detail, 'total_quantity_decimal': 2.05}, ... ]}
    final_loading_sheets = {}
    for vehicle, product_item_counts in vehicle_loading_sheets_agg.items():
        
        loading_sheet_items = []
        for product_detail, total_items in product_item_counts.items():
            # Use the ProductDetail's own helper method to convert the total item count back
            # into the 'MasterUnits.IndividualItems' decimal format.
            total_quantity_decimal = product_detail._get_decimal_from_items(total_items)
            
            loading_sheet_items.append({
                'product': product_detail,
                'total_quantity_decimal': total_quantity_decimal,
            })

        # Sort items by product name for a clean printout
        final_loading_sheets[vehicle] = sorted(loading_sheet_items, key=lambda x: x['product'].product_base.name)

    # --- Total revenue credit and online and receiveable per vehicle ---
    vehicle_total_revenue = defaultdict(lambda: Decimal('0.00'))
    vehicle_total_discount = defaultdict(lambda: Decimal('0.00'))
    vehicle_total_credit = defaultdict(lambda: Decimal('0.00'))
    vehicle_total_online = defaultdict(lambda: Decimal('0.00'))
    vehicle_remaining_amount = defaultdict(lambda: Decimal('0.00'))

    for vehicle, group in groupby(pending_transactions, key=attrgetter('assigned_vehicle')):
        group = list(group)  # Convert groupby iterator to list to iterate multiple times
        vehicle_total_revenue[vehicle.pk] = sum(tx.grand_total_revenue for tx in group)
        vehicle_total_discount[vehicle.pk] = sum(tx.total_discount_amount or Decimal('0.00') for tx in group)
        vehicle_total_credit[vehicle.pk] = sum(tx.amount_on_credit or Decimal('0.00') for tx in group)
        vehicle_total_online[vehicle.pk] = sum(tx.amount_paid_online or Decimal('0.00') for tx in group)
        vehicle_remaining_amount[vehicle.pk] = (vehicle_total_revenue[vehicle.pk]- vehicle_total_credit[vehicle.pk]- vehicle_total_online[vehicle.pk])
    context = {
        'pending_transactions': pending_transactions,
        'loading_sheets': final_loading_sheets,
        'vehicle_total_revenue': vehicle_total_revenue,
        'vehicle_total_discount': vehicle_total_discount,
        'vehicle_total_credit': vehicle_total_credit,
        'vehicle_total_online': vehicle_total_online,
        'vehicle_remaining_amount': vehicle_remaining_amount,
        'no_of_invoices': no_of_invoices,
    }
    return render(request, 'stock/pending_deliveries.html', context)





@login_required
def process_delivery_return_view(request, sale_pk):
    sales_transaction = get_object_or_404(
        SalesTransaction, 
        pk=sale_pk, 
        user=request.user, 
        status__in=['PENDING_DELIVERY', 'PARTIALLY_RETURNED']
    )
    items_queryset = SalesTransactionItem.objects.filter(transaction=sales_transaction)

    if request.method == 'POST':
        return_formset = SalesTransactionItemReturnFormSet(request.POST, queryset=items_queryset)
        delivery_form = ProcessDeliveryForm(request.POST, instance=sales_transaction)

        if return_formset.is_valid() and delivery_form.is_valid():
            mark_as_done = 'mark_as_done' in request.POST
            try:
                if mark_as_done:
                    # Save updated return quantities and demand edits
                    final_cost_of_tx = Decimal('0.00')
                    for form_item in return_formset:
                        item_instance = form_item.save(commit=False)
                        product = item_instance.product_detail_snapshot
                        returned_qty = form_item.cleaned_data.get('returned_quantity_decimal', Decimal('0.00'))
                        increased_demand = form_item.cleaned_data.get('increased_demand') or Decimal('0.00')

                        # Adjust quantity sold for increased demand
                        if increased_demand > 0:
                            item_instance.returned_quantity_decimal = returned_qty
                            item_instance.increased_demand = increased_demand

                        item_instance.returned_quantity_decimal = returned_qty
                        item_instance.save()

                        dispatched_items = product._get_items_from_decimal(item_instance.quantity_sold_decimal)
                        returned_items = product._get_items_from_decimal(returned_qty)
                        increased_items = product._get_items_from_decimal(increased_demand)

                        net_items = dispatched_items + increased_items - returned_items
                        final_cost_of_tx += net_items * item_instance.selling_price_per_item


                    sales_transaction.total_discount_amount = delivery_form.cleaned_data['total_discount_amount']
                    sales_transaction.payment_type = delivery_form.cleaned_data['payment_type']
                    sales_transaction.amount_paid_cash = delivery_form.cleaned_data['amount_paid_cash']
                    sales_transaction.amount_paid_online = delivery_form.cleaned_data['amount_paid_online']
                    sales_transaction.amount_on_credit = delivery_form.cleaned_data['amount_on_credit']
                    discount = sales_transaction.total_discount_amount or Decimal('0.00')
                    expected_total = final_cost_of_tx - discount
                    paid_total = (sales_transaction.amount_on_credit or Decimal('0.00')) + \
                                (sales_transaction.amount_paid_online or Decimal('0.00')) + \
                                (sales_transaction.amount_paid_cash or Decimal('0.00'))

                    if paid_total > expected_total:
                        raise forms.ValidationError("Total paid amount exceeds final total after discount.")
                    elif paid_total < expected_total:
                        raise forms.ValidationError("Total paid amount is less than final total after discount.")
                    sales_transaction.grand_total_revenue = final_cost_of_tx - discount
                    sales_transaction.is_ready_for_processing = True
                    sales_transaction.status = 'PENDING_DELIVERY'
                    sales_transaction.save()
                    messages.success(request, "Marked as done. Will be processed with your changes.")
                    
                    return redirect('stock:pending_deliveries')
                    # mark as done is completed and controll is returned ......

                # Full processing
                with transaction.atomic():
                    for form_item in return_formset:
                        item_instance = form_item.instance
                        product_detail = item_instance.product_detail_snapshot

                        # Process returned quantity change
                        if form_item.has_changed() and 'returned_quantity_decimal' in form_item.changed_data:
                            original_returned = item_instance.returned_quantity_decimal
                            new_returned = form_item.cleaned_data['returned_quantity_decimal']
                            stock_diff = new_returned - original_returned
                            if stock_diff > 0:
                                product_detail.increase_stock(stock_diff)
                            elif stock_diff < 0:
                                product_detail.decrease_stock(abs(stock_diff))

                        # Process increased demand
                        increased_demand = form_item.cleaned_data.get('increased_demand') or Decimal('0.00')
                        if increased_demand > 0:
                            original_qty = item_instance.quantity_sold_decimal or Decimal('0.00')
                            total_items = product_detail._get_items_from_decimal(original_qty) + product_detail._get_items_from_decimal(increased_demand)
                            updated_qty = product_detail._get_decimal_from_items(total_items)
                            item_instance.quantity_sold_decimal = updated_qty
                            product_detail.decrease_stock(increased_demand)
                            item_instance.save()

                    return_formset.save()

                    # Recalculate grand total
                    new_discount = delivery_form.cleaned_data.get('total_discount_amount') or Decimal('0.00')
                    sales_transaction.update_grand_totals(new_discount_amount=new_discount)
                    sales_transaction.refresh_from_db()
                    final_grand_total = sales_transaction.grand_total_revenue

                    # Payment processing
                    payment_type = delivery_form.cleaned_data['payment_type']
                    cash = delivery_form.cleaned_data.get('amount_paid_cash') or Decimal('0.00')
                    online = delivery_form.cleaned_data.get('amount_paid_online') or Decimal('0.00')
                    credit = delivery_form.cleaned_data.get('amount_on_credit') or Decimal('0.00')

                    if payment_type == 'SPLIT':
                        total_split = (cash + online + credit).quantize(Decimal('0.01'))
                        if total_split != final_grand_total.quantize(Decimal('0.01')):
                            raise forms.ValidationError(f"Split payments (Rs {total_split}) do not match Grand Total (Rs {final_grand_total}).")
                        sales_transaction.amount_paid_cash = cash
                        sales_transaction.amount_paid_online = online
                        sales_transaction.amount_on_credit = credit
                        sales_transaction.notes = (
                            f"Split Payment: Cash = Rs {cash:.2f}, "
                            f"Online = Rs {online:.2f}, "
                            f"Credit = Rs {credit:.2f}"
                        )
                    else:
                        sales_transaction.amount_paid_cash = final_grand_total if payment_type == 'CASH' else Decimal('0.00')
                        sales_transaction.amount_paid_online = final_grand_total if payment_type == 'ONLINE' else Decimal('0.00')
                        sales_transaction.amount_on_credit = final_grand_total if payment_type == 'CREDIT' else Decimal('0.00')

                    sales_transaction.payment_type = payment_type

                    # Update status
                    total_dispatched = sum(item.dispatched_individual_items_count for item in sales_transaction.items.all())
                    total_returned = sum(item.returned_individual_items_count for item in sales_transaction.items.all())

                    if total_returned == 0:
                        sales_transaction.status = 'COMPLETED'
                    elif total_returned >= total_dispatched:
                        sales_transaction.status = 'FULLY_RETURNED'
                    else:
                        sales_transaction.status = 'PARTIALLY_RETURNED'

                    sales_transaction.save()

                    # Financial Ledger
                    existing_credit = ShopFinancialTransaction.objects.filter(source_sale=sales_transaction).first()
                    if sales_transaction.amount_on_credit > 0 and sales_transaction.customer_shop:
                        if existing_credit:
                            existing_credit.debit_amount = sales_transaction.amount_on_credit
                            existing_credit.save()
                        else:
                            ShopFinancialTransaction.objects.create(
                                shop=sales_transaction.customer_shop, user=request.user,
                                source_sale=sales_transaction, transaction_type='CREDIT_SALE',
                                debit_amount=sales_transaction.amount_on_credit
                            )
                    elif existing_credit:
                        existing_credit.delete()

                    messages.success(request, f"Delivery for Transaction #{sales_transaction.pk} processed successfully.")
                    return redirect('stock:pending_deliveries')

            except forms.ValidationError as e:
                delivery_form.add_error(None, e)
                messages.error(request, str(e.message))
            except Exception as e:
                messages.error(request, f"A critical error occurred: {str(e)}")
        else:
            messages.error(request, "Please correct the errors shown in the forms below.")
    else:
        return_formset = SalesTransactionItemReturnFormSet(queryset=items_queryset)
        delivery_form = ProcessDeliveryForm(instance=sales_transaction)

    return render(request, 'stock/process_delivery_return.html', {
        'return_formset': return_formset,
        'delivery_form': delivery_form,
        'transaction': sales_transaction,
    })



@login_required
@transaction.atomic
def process_all_pending_for_vehicle(request, vehicle_id):
    vehicle = get_object_or_404(Vehicle, pk=vehicle_id)

    transactions = SalesTransaction.objects.filter(
        assigned_vehicle=vehicle,
        user=request.user,
        status='PENDING_DELIVERY'
    ).prefetch_related('items__product_detail_snapshot', 'customer_shop')

    if not transactions.exists():
        messages.warning(request, f"No pending deliveries found for vehicle {vehicle.vehicle_number}.")
        return redirect('stock:pending_deliveries')

    try:
        for tx in transactions:
            final_total = Decimal('0.00')
            for item in tx.items.all():
                product = item.product_detail_snapshot
                selling_price = item.selling_price_per_item or Decimal('0.00')

                dispatched_items = product._get_items_from_decimal(item.quantity_sold_decimal or Decimal('0.00'))
                returned_items = product._get_items_from_decimal(item.returned_quantity_decimal or Decimal('0.00'))
                increased_items = product._get_items_from_decimal(item.increased_demand or Decimal('0.00'))

                # Apply stock changes
                if returned_items > 0:
                    product.increase_stock(product._get_decimal_from_items(returned_items))

                if increased_items > 0:
                    product.decrease_stock(product._get_decimal_from_items(increased_items))

                # Update sold quantity (original + increased)
                new_total_items = dispatched_items + increased_items
                item.quantity_sold_decimal = product._get_decimal_from_items(new_total_items)
                item.save()

                net_items = new_total_items - returned_items
                net_items = max(net_items, 0)
                final_total += selling_price * net_items

            # Apply discount
            discount = tx.total_discount_amount or Decimal('0.00')
            final_total -= discount

            tx.grand_total_revenue = final_total

            # Process payment fields
            if tx.payment_type == 'SPLIT':
                cash = tx.amount_paid_cash or Decimal('0.00')
                online = tx.amount_paid_online or Decimal('0.00')
                credit = tx.amount_on_credit or Decimal('0.00')
                tx.notes = f"Split Payment: Cash = Rs {cash:.2f}, Online = Rs {online:.2f}, Credit = Rs {credit:.2f}"
            else:
                cash = final_total if tx.payment_type == 'CASH' else Decimal('0.00')
                online = final_total if tx.payment_type == 'ONLINE' else Decimal('0.00')
                credit = final_total if tx.payment_type == 'CREDIT' else Decimal('0.00')
                tx.notes = ""

            tx.amount_paid_cash = cash
            tx.amount_paid_online = online
            tx.amount_on_credit = credit
            tx.status = 'COMPLETED'
            tx.is_ready_for_processing = False  # Reset manual flag
            tx.save()

            # Update financial ledger
            existing_ledger = ShopFinancialTransaction.objects.filter(source_sale=tx).first()
            if credit > 0 and tx.customer_shop:
                if existing_ledger:
                    existing_ledger.debit_amount = credit
                    existing_ledger.save()
                else:
                    ShopFinancialTransaction.objects.create(
                        shop=tx.customer_shop,
                        user=request.user,
                        source_sale=tx,
                        transaction_type='CREDIT_SALE',
                        debit_amount=credit
                    )
            elif existing_ledger:
                existing_ledger.delete()

        messages.success(request, f"All pending deliveries for vehicle {vehicle.vehicle_number} processed successfully.")

    except Exception as e:
        messages.error(request, f"A critical error occurred: {str(e)}")

    return redirect('stock:pending_deliveries')


@login_required
@admin_mode_required
def sales_report_view(request):
    today = timezone.localdate()
    # ... (date range setup: start_of_today_dt, end_of_today_dt, etc. - as before) ...
    # Ensure these are timezone-aware datetimes for comparison with transaction_time
    start_of_today  = timezone.make_aware(timezone.datetime.combine(today, timezone.datetime.min.time()))
    end_of_today  = timezone.make_aware(timezone.datetime.combine(today, timezone.datetime.max.time()))

    # ... similarly for week, month, year ...
    start_of_week = start_of_today - timedelta(days=today.weekday())
    end_of_week = start_of_week + timedelta(days=6, hours=23, minutes=59, seconds=59)

    start_of_month = start_of_today.replace(day=1)
    next_month = start_of_month.replace(day=28) + timedelta(days=4)
    end_of_month = next_month - timedelta(days=next_month.day)
    end_of_month = timezone.make_aware(timezone.datetime.combine(end_of_month.date(), timezone.datetime.max.time()))
    
    start_of_year = start_of_today.replace(month=1, day=1)
    end_of_year = start_of_year.replace(year=start_of_year.year + 1) - timedelta(days=1)
    end_of_year = timezone.make_aware(timezone.datetime.combine(end_of_year.date(), timezone.datetime.max.time()))


    def get_period_stats(start_dt, end_dt):
        """
        Helper function to get sales and expense stats for a given period.
        """
        transactions = SalesTransaction.objects.filter(
            user=request.user, 
            transaction_time__range=(start_dt, end_dt)
        )
        expenses = Expense.objects.filter(
            user=request.user, 
            expense_date__range=(start_dt, end_dt)
        )
        claims = Claim.objects.filter(
            user=request.user,
            status='COMPLETED',
            claim_date__range=(start_dt, end_dt)
        )
        stock_claim_loss = sum(claim.value_of_items_given for claim in claims) or Decimal('0.00')
        
        total_revenue = sum(tx.grand_total_revenue for tx in transactions)
        total_profit = sum(tx.calculated_grand_profit for tx in transactions)
        total_expense = expenses.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
                
        net_profit = total_profit - total_expense - stock_claim_loss
        
        return {
            'transactions_count': transactions.count(),
            'total_grand_revenue': total_revenue,
            'total_grand_profit': total_profit,
            'total_expense': total_expense,
            'stock_claim_loss': stock_claim_loss,
            'net_profit': net_profit,
        }

    stats_today = get_period_stats(start_of_today, end_of_today)
    stats_this_week = get_period_stats(start_of_week, end_of_week)
    stats_this_month = get_period_stats(start_of_month, end_of_month)
    stats_this_year = get_period_stats(start_of_year, end_of_year)

    # --- Data for Charts (using SalesTransaction) ---
    daily_labels = []
    daily_revenue_data = []
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        start = timezone.make_aware(timezone.datetime.combine(day, timezone.datetime.min.time()))
        end = timezone.make_aware(timezone.datetime.combine(day, timezone.datetime.max.time()))

        tx_on_day = SalesTransaction.objects.filter(user=request.user, transaction_time__range=(start, end))
        daily_revenue = sum(tx.grand_total_revenue for tx in tx_on_day) or Decimal('0.00')

        daily_labels.append(day.strftime("%a")) # Short day name e.g., "Mon"
        daily_revenue_data.append(float(daily_revenue))



    # ... Monthly chart data logic ...
    monthly_labels = []
    monthly_revenue_data = []
    for i in range(5, -1, -1):
        # ... (logic for first_day_of_target_month, last_day_of_target_month as before) ...

        target_month_date = today # Start from today for current month calculation
        
        for _ in range(i):
            first_day_of_prev_month = target_month_date.replace(day=1) - timedelta(days=1)
            target_month_date = first_day_of_prev_month
        start_of_target_month = target_month_date.replace(day=1)
        next_m = start_of_target_month.replace(day=28) + timedelta(days=4)
        end_of_target_month = next_m - timedelta(days=next_m.day)
        
        start_dt = timezone.make_aware(timezone.datetime.combine(start_of_target_month, timezone.datetime.min.time()))
        end_dt = timezone.make_aware(timezone.datetime.combine(end_of_target_month, timezone.datetime.max.time()))
        
        tx_in_month = SalesTransaction.objects.filter(user=request.user, transaction_time__range=(start_dt, end_dt))
        monthly_revenue = sum(tx.grand_total_revenue for tx in tx_in_month) or Decimal('0.00')
        
        monthly_labels.append(start_of_target_month.strftime("%b %Y"))
        monthly_revenue_data.append(float(monthly_revenue))
        
    context = {
        'stats_today': stats_today,
        'stats_this_week': stats_this_week,
        'stats_this_month': stats_this_month,
        'stats_this_year': stats_this_year,
        'report_generation_time': timezone.now(),
        'daily_chart_labels': json.dumps(daily_labels),
        'daily_chart_revenue': json.dumps(daily_revenue_data),
        'monthly_chart_labels': json.dumps(monthly_labels),
        'monthly_chart_revenue': json.dumps(monthly_revenue_data),
    }
    return render(request, 'stock/sales_report.html', context)


# Recepit View ...................................

@login_required
def sale_receipt_view(request, sale_pk): # sale_pk is now the PK of SalesTransaction
    # Fetch the SalesTransaction, ensuring it belongs to the current user
    # Pre-fetch related items and their product details for efficiency
    sales_transaction = get_object_or_404(
        SalesTransaction.objects.select_related(
            'user', 
            'customer_shop', 
            'assigned_vehicle'
        ).prefetch_related(
            'items__product_detail_snapshot__product_base' # Prefetch items and their nested details
        ),
        pk=sale_pk,
        user=request.user # Or adjust if admins can print any receipt
    )
    admin_profile = AdminProfile.objects.only('company_name','company_address','company_phone').filter(user = request.user)
    for profile in admin_profile:
        company_name = profile.company_name
        company_address = profile.company_address
        company_phone = profile.company_phone

    context = {
        'transaction': sales_transaction, # Pass the whole transaction object
        'company_name': company_name,
        'company_address': company_address,
        'company_phone': company_phone,
    }
    return render(request, 'stock/sale_receipt.html', context)

# vehical views ............................................................................................................
@login_required
def manage_vehicles_view(request):
    add_form_in_modal_has_errors = False 
    add_vehicle_form = VehicleForm() # For the "Add New Vehicle" modal

    if request.method == 'POST':
        # This assumes POST to this view is for creating new vehicle.
        # Update actions will POST to vehicle_update_action_view.
        filled_add_form = VehicleForm(request.POST)
        if filled_add_form.is_valid():
            vehicle_instance = filled_add_form.save(commit=False)
            vehicle_instance.user = request.user
            vehicle_instance.save()
            messages.success(request, f"Vehicle '{vehicle_instance.vehicle_number}' added successfully!")
            return redirect('stock:manage_vehicles')
        else:
            messages.error(request, "Error adding vehicle. Please correct the form.")
            add_form_in_modal_has_errors = True
            add_vehicle_form = filled_add_form # Pass form with errors back to "Add" modal
    
    vehicles_list = Vehicle.objects.filter(user=request.user).order_by('-created_at')
    context = {
        'add_form': add_vehicle_form, # Use a specific name for the add form
        'vehicles': vehicles_list,
        'add_form_in_modal_has_errors': add_form_in_modal_has_errors,
    }
    return render(request, 'stock/add_vehicles.html', context)


@login_required
def vehicle_update_action_view(request, vehicle_pk):
    instance = get_object_or_404(Vehicle, pk=vehicle_pk, user=request.user)
    if request.method == 'POST':
        form = VehicleForm(request.POST, instance=instance)
        if form.is_valid():
            # The user field is already set on instance and is not part of VehicleForm fields by default
            # If you needed to change the user, you'd handle it explicitly.
            # For now, assume only user who created can update.
            # vehicle_to_update = form.save(commit=False)
            # vehicle_to_update.user = request.user # Ensure user is still correct if it was part of form
            form.save() # This updates the instance
            messages.success(request, f"Vehicle '{instance.vehicle_number}' updated successfully!")
            return redirect('stock:manage_vehicles')
        else:
            # This is tricky for a modal on another page without full page reload for the update form.
            # Ideally, the update modal form would submit via AJAX to this view.
            # If it fails, this view returns JSON errors, and JS updates the modal.
            # For a non-AJAX POST from the modal:
            # We can't easily re-render the list page with the update modal open and errors.
            # So, we redirect with a generic error. The user would have to re-open the modal.
            error_message = "Failed to update vehicle. Please correct the errors: "
            for field, errors in form.errors.items():
                error_message += f"{field}: {', '.join(errors)} "
            messages.error(request, error_message.strip())
            # To pass form errors back to modal on list page is complex without AJAX for the update itself.
            # Simplest is redirect. User will lose their unsaved changes in modal.
            return redirect('stock:manage_vehicles') 
    else:
        # This view is primarily for POST. A GET request isn't typical for modal updates.
        return redirect('stock:manage_vehicles')
    
@login_required
def ajax_get_vehicle_data(request, vehicle_pk):
    try:
        vehicle = get_object_or_404(Vehicle, pk=vehicle_pk, user=request.user)
        data = {
            'pk': vehicle.pk,
            'vehicle_number': vehicle.vehicle_number,
            'vehicle_type': vehicle.vehicle_type,
            'driver_name': vehicle.driver_name or "",
            'driver_phone': vehicle.driver_phone or "",
            'capacity_kg': str(vehicle.capacity_kg) if vehicle.capacity_kg is not None else "",
            'notes': vehicle.notes or "",
            'is_active': vehicle.is_active,
        }
        return JsonResponse({'success': True, 'data': data})
    except Vehicle.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Vehicle not found or not authorized.'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    




@login_required
def delete_vehicle_action_view(request, vehicle_pk):
    # Fetch the vehicle, ensuring it belongs to the current user
    vehicle_to_delete = get_object_or_404(Vehicle, pk=vehicle_pk, user=request.user)

    if request.method == 'POST': # This will be from the modal's form
        vehicle_number = vehicle_to_delete.vehicle_number
        try:
            vehicle_to_delete.delete()
            messages.success(request, f"Vehicle '{vehicle_number}' has been deleted successfully.")
        except Exception as e:
            messages.error(request, f"Error deleting vehicle '{vehicle_number}': {str(e)}")
        return redirect('stock:manage_vehicles') # Redirect back to the vehicle list
    else:
        # Direct GET access to this URL is not intended with the modal approach
        messages.warning(request, "Invalid request for vehicle deletion.")
        return redirect('stock:manage_vehicles')
    



# shop views.................................................................................................
@login_required
def manage_shops_view(request):
    add_form_in_modal_has_errors = False 
    add_shop_form = ShopForm() # For the "Add New Shop" modal
    query = request.GET.get('q','').strip()
    shops_queryset = Shop.objects.filter(user=request.user).order_by('name')
    if query:
        shops_queryset = shops_queryset.filter(
            Q(name__icontains=query) |
            Q(location_address__icontains=query) |
            Q(contact_person__icontains=query) |
            Q(contact_phone__icontains=query)
        )
    paginator = Paginator(shops_queryset, 100)
    page_number = request.GET.get('page')
    try:
        shops_page = paginator.get_page(page_number)
    except PageNotAnInteger:
        shops_page = paginator.page(1)
    except EmptyPage:
        shops_page = paginator.page(paginator.num_pages)

    if request.method == 'POST':
        filled_add_form = ShopForm(request.POST)
        if filled_add_form.is_valid():
            shop_instance = filled_add_form.save(commit=False)
            shop_instance.user = request.user
            shop_instance.save()
            messages.success(request, f"Shop '{shop_instance.name}' added successfully!")
            return redirect('stock:manage_shops')
        else:
            messages.error(request, "Error adding shop. Please correct the form.")
            add_form_in_modal_has_errors = True
            add_shop_form = filled_add_form # Pass form with errors back to "Add" modal
    context = {
        'add_form': add_shop_form, # Use a specific name for the add form
        'shops': shops_page,
        'page_obj': shops_page,
        'add_form_in_modal_has_errors': add_form_in_modal_has_errors,
        'query': query,
    }
    return render(request, 'stock/add_shops.html', context)

@login_required
def delete_shop_action_view(request, shop_pk):
    shop_to_delete = get_object_or_404(Shop, pk=shop_pk, user=request.user)

    if request.method == 'POST': # From modal confirmation
        shop_name = shop_to_delete.name
        try:
            shop_to_delete.delete()
            messages.success(request, f"Shop '{shop_name}' has been deleted successfully.")
        except Exception as e:
            # Handle potential ProtectedError if shops are linked to other models that prevent deletion
            messages.error(request, f"Error deleting shop '{shop_name}': {str(e)}. It might be linked to other records.")
        return redirect('stock:manage_shops')
    else:
        messages.warning(request, "Invalid request for shop deletion.")
        return redirect('stock:manage_shops')
    

@login_required
def shop_update_action_view(request, shop_pk):
    instance = get_object_or_404(Shop, pk=shop_pk, user=request.user)
    if request.method == 'POST':
        form = ShopForm(request.POST, instance=instance)
        if form.is_valid():
            form.save() # This updates the instance
            messages.success(request, f"Shop '{instance.name}' updated successfully!")
            return redirect('stock:manage_shops')
        else:
            error_message = "Failed to update shop. Please correct the errors: "
            for field, errors in form.errors.items(): # Concatenate all errors
                error_message += f"Field '{field}': {', '.join(errors)} "
            messages.error(request, error_message.strip())
            # Redirect back, user will have to click edit again.
            # For better UX, the modal form would POST via AJAX to this view.
            return redirect('stock:manage_shops') 
    else:
        # This view is primarily for POST from the modal.
        return redirect('stock:manage_shops')
    

@login_required
def ajax_get_shop_data(request, shop_pk):
    try:
        shop = get_object_or_404(Shop, pk=shop_pk, user=request.user)
        data = {
            'pk': shop.pk,
            'name': shop.name,
            'location_address': shop.location_address or "",
            'contact_person': shop.contact_person or "",
            'contact_phone': shop.contact_phone or "",
            'email': shop.email or "",
            'notes': shop.notes or "",
            'is_active': shop.is_active,
        }
        return JsonResponse({'success': True, 'data': data})
    except Shop.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Shop not found or not authorized.'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)



@login_required
def list_shops_for_sales_view(request):
    # This view should be fine if it only lists Shop instances.
    # The links it generates will point to shop_purchase_history_view.
    shops = Shop.objects.filter(user=request.user, is_active=True).order_by('name')
    
    # Filter shops to only include those with transactions having grand_total_revenue > 0
    shops_with_revenue = []
    for shop in shops:
        shop_transactions = SalesTransaction.objects.filter(
            customer_shop=shop,
            user=request.user 
        ).select_related(
            'assigned_vehicle'
        ).prefetch_related(
            'items__product_detail_snapshot__product_base'
        ).order_by('-transaction_time')
        
        # Check if this shop has any transactions with revenue > 0
        total_revenue = sum(tx.grand_total_revenue for tx in shop_transactions)
        if total_revenue > 0:
            shops_with_revenue.append(shop)

    context = {'shops': shops_with_revenue}
    return render(request, 'stock/list_shops_for_sales.html', context)

@login_required
def shop_purchase_history_view(request, shop_pk):
    shop = get_object_or_404(Shop, pk=shop_pk) # Potentially add user=request.user if shops are user-scoped

    # Fetch SalesTransaction records associated with this shop and processed by the current user
    shop_transactions = SalesTransaction.objects.filter(
        customer_shop=shop,
        user=request.user 
    ).select_related(
        'assigned_vehicle' # No need to select_related product_detail here, items are separate
    ).prefetch_related(
        'items__product_detail_snapshot__product_base'
    ).order_by('-transaction_time')

    # Recalculate totals based on the new structure
    shop_total_revenue = sum(tx.grand_total_revenue for tx in shop_transactions)
    shop_total_profit = sum(tx.calculated_grand_profit for tx in shop_transactions)

    context = {
        'shop': shop,
        'shop_transactions': shop_transactions, # Changed context variable name
        'shop_total_revenue': shop_total_revenue,
        'shop_total_profit': shop_total_profit,
    }
    return render(request, 'stock/shop_purchase_history.html', context)


@login_required
def sales_by_group_hub_view(request):
    """
    Displays a navigation hub with cards for each vehicle and one for the "Store".
    """
    # Get all active vehicles and annotate them with their sales count for display.
    vehicles = Vehicle.objects.filter(
        user=request.user, 
        is_active=True
    ).annotate(
        sales_count=Count('assigned_sales_transactions')
    ).order_by('vehicle_number')

    # Get the count of sales not assigned to any vehicle.
    store_sales_count = SalesTransaction.objects.filter(
        user=request.user,
        assigned_vehicle__isnull=True
    ).count()

    context = {
        'vehicles': vehicles,
        'store_sales_count': store_sales_count,
    }
    return render(request, 'stock/sales_by_group_hub.html', context)





@login_required
def performance_summary_hub_view(request):
    """
    Displays a navigation hub with cards for each vehicle and one for the "Store"
    to link to the performance summary page.
    """
    vehicles = Vehicle.objects.filter(user=request.user, is_active=True).order_by('vehicle_number')
    context = {
        'vehicles': vehicles,
    }
    return render(request, 'stock/performance_summary_hub.html', context)


@login_required
def group_performance_summary_view(request, vehicle_pk=None):
    """
    DEFINITIVE VERSION 3: This view calculates figures dynamically. It correctly
    aggregates quantities in Python and formats the output to match the
    existing template's expectations without requiring HTML changes.
    """
    user = request.user
    
    if vehicle_pk:
        grouping_object = get_object_or_404(Vehicle, pk=vehicle_pk)
        grouping_name = f"Performance Summary for: {grouping_object.vehicle_number}"
        base_items_query = SalesTransactionItem.objects.filter(
            transaction__user=user, transaction__assigned_vehicle=grouping_object
        )
    else:
        grouping_object = None
        grouping_name = "Performance Summary for: Store"
        base_items_query = SalesTransactionItem.objects.filter(
            transaction__user=user, transaction__assigned_vehicle__isnull=True
        )

    # Pre-fetch related data for efficiency
    base_items_query = base_items_query.select_related(
        'transaction', 'product_detail_snapshot__product_base'
    )
    
    # --- Daily Summary Calculation ---
    days_with_sales = base_items_query.annotate(
        day=TruncDate('transaction__transaction_time')
    ).values_list('day', flat=True).distinct().order_by('-day')
    
    daily_summary = []
    for day in days_with_sales:
        items_on_day = base_items_query.filter(transaction__transaction_time__date=day)
        
        # Calculate net revenue in Python for accuracy
        net_revenue_on_day = sum(item.gross_line_subtotal for item in items_on_day)

        # --- PYTHON-BASED NET QUANTITY CALCULATION (THE FIX) ---
        # 1. Group items by their packaging and sum their individual item counts.
        quantity_agg = defaultdict(int)
        # Store a reference object for each group to access its properties later
        ref_objects = {}
        for item in items_on_day:
            pd_snapshot = item.product_detail_snapshot
            # Use the product detail's PK as a simple, hashable key
            key = pd_snapshot.pk
            quantity_agg[key] += item.actual_sold_individual_items_count
            if key not in ref_objects:
                ref_objects[key] = pd_snapshot

        # 2. Convert the summed individual item counts back to the decimal format.
        quantity_breakdown = []
        for pk, total_items in quantity_agg.items():
            if total_items > 0:
                # Get the reference ProductDetail object for this group
                ref_product_detail = ref_objects[pk]
                
                # Call the method from the correct model (ProductDetail)
                net_quantity_decimal = ref_product_detail._get_decimal_from_items(total_items)

                # Format the dictionary EXACTLY as the template expects
                quantity_breakdown.append({
                    'net_quantity': net_quantity_decimal,
                    'product_detail_snapshot__quantity_in_packing': ref_product_detail.quantity_in_packing,
                    'product_detail_snapshot__unit_of_measure': ref_product_detail.unit_of_measure,
                })
        
        daily_summary.append({
            'day': day,
            'total_revenue': net_revenue_on_day,
            'quantity_breakdown': sorted(quantity_breakdown, key=lambda x: x['net_quantity'], reverse=True)
        })

    # --- Monthly Summary Calculation (with the same correct logic) ---
    months_with_sales = base_items_query.annotate(
        month=TruncMonth('transaction__transaction_time')
    ).values_list('month', flat=True).distinct().order_by('-month')

    monthly_summary = []
    for month in months_with_sales:
        items_in_month = base_items_query.filter(
            transaction__transaction_time__year=month.year,
            transaction__transaction_time__month=month.month
        )
        
        net_revenue_in_month = sum(item.gross_line_subtotal for item in items_in_month)
        
        # --- REPEAT THE SAME PYTHON-BASED NET QUANTITY CALCULATION ---
        quantity_agg = defaultdict(int)
        ref_objects = {}
        for item in items_in_month:
            pd_snapshot = item.product_detail_snapshot
            key = pd_snapshot.pk
            quantity_agg[key] += item.actual_sold_individual_items_count
            if key not in ref_objects:
                ref_objects[key] = pd_snapshot
        
        quantity_breakdown = []
        for pk, total_items in quantity_agg.items():
            if total_items > 0:
                ref_product_detail = ref_objects[pk]
                net_quantity_decimal = ref_product_detail._get_decimal_from_items(total_items)
                quantity_breakdown.append({
                    'net_quantity': net_quantity_decimal,
                    'product_detail_snapshot__quantity_in_packing': ref_product_detail.quantity_in_packing,
                    'product_detail_snapshot__unit_of_measure': ref_product_detail.unit_of_measure,
                })

        monthly_summary.append({
            'month': month,
            'total_revenue': net_revenue_in_month,
            'quantity_breakdown': sorted(quantity_breakdown, key=lambda x: x['net_quantity'], reverse=True)
        })

    context = {
        'grouping_name': grouping_name,
        'daily_summary': daily_summary,
        'monthly_summary': monthly_summary,
    }
    return render(request, 'stock/group_performance_summary.html', context)


@login_required
def sales_group_details_view(request, vehicle_pk=None):
    """
    Displays a filtered list of sales transactions for a specific group
    (either a vehicle or the store).
    """
    user = request.user
    
    # Start with the base query for all sales for the user.
    base_query = SalesTransaction.objects.filter(user=user).select_related(
        'customer_shop', 'assigned_vehicle'
    ).prefetch_related('items__product_detail_snapshot__product_base')

    # Filter the query based on the group selected.
    if vehicle_pk:
        # User selected a specific vehicle.
        grouping_object = get_object_or_404(Vehicle, pk=vehicle_pk, user=user)
        grouping_name = f"Sales for Vehicle: {grouping_object.vehicle_number}"
        transactions_list = base_query.filter(assigned_vehicle=grouping_object)
    else:
        # User selected the "Store".
        grouping_object = None
        grouping_name = "Store Sales (No Vehicle)"
        transactions_list = base_query.filter(assigned_vehicle__isnull=True)

    # --- Date Filtering (Optional but recommended) ---
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date:
        transactions_list = transactions_list.filter(transaction_time__date__gte=start_date)
    if end_date:
        transactions_list = transactions_list.filter(transaction_time__date__lte=end_date)
    
    # Final ordering
    transactions_list = transactions_list.order_by('-transaction_time')

    # --- Pagination ---
    paginator = Paginator(transactions_list, 50) # 50 sales per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'grouping_name': grouping_name,
        'page_obj': page_obj,
        'transactions': page_obj.object_list,
    }
    # We can reuse the main transaction list template.
    return render(request, 'stock/all_transactions_list.html', context)